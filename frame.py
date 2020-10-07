import telebot
import openpyxl
from telebot import types
import os
from datetime import datetime
import threading
import time
import urllib

def srok(j):
    if j == 2:
        q = 1
        p = 7
    elif j == 3 or j == 4:
        q = 12
        p = 14
    else:
        q = 6
        p = 7
    return q, p

def organiza(user_id):
    obj_ok = openpyxl.load_workbook('Исход.xlsx')
    sheet_obj_ok = obj_ok['Данные']
    quantity = sheet_obj_ok.max_row
    bot.send_message(user_id, 'Организации которые есть в списке:')
    for i in range(quantity):
        if i + 1 == 1:
            continue
        else:
            data = sheet_obj_ok.cell(row=i + 1, column=1).value
            bot.send_message(user_id, '*' + data + '*', parse_mode="markdown")

def find_org(user_id, tex):
    obj_ok = openpyxl.load_workbook('Исход.xlsx')
    sheet_obj_ok = obj_ok['Данные']
    quantity = sheet_obj_ok.max_row
    for i in range(quantity):
        if i + 1 == 1:
            continue
        else:
            if str(sheet_obj_ok.cell(row=i + 1, column=1).value).strip().lower() == tex:
                for j in range(6):
                    if j + 1 == 1:
                        continue
                    else:
                        if sheet_obj_ok.cell(row=i + 1, column=j + 1).value is None:
                            data = 'сведения отсутствуют'
                        else:
                            dataa = str(sheet_obj_ok.cell(row=i + 1, column=j + 1).value).split()[0].split('-')
                            data = dataa[-1] + '.' + dataa[1] + '.' + dataa[0]
                        dok = str(sheet_obj_ok.cell(row=1, column=j + 1).value)
                        bot.send_message(user_id, 'Дата выдачи *' + dok + '* является *' + data + '*',
                                         parse_mode="markdown")
                break

if not os.path.exists('Исход.xlsx'):
    new = openpyxl.Workbook()  # создание нового экселя
    new.save('Исход.xlsx')  # его сохранение
    obj = openpyxl.load_workbook('Исход.xlsx')
    sheet_obj = obj.active
    sheet_obj.title = 'Данные'
    sheet_obj['A1'] = 'Организация'
    sheet_obj['B1'] = 'СРО'
    sheet_obj['C1'] = 'ЭЦП'
    sheet_obj['D1'] = 'Протокол крупной сделки'
    sheet_obj['E1'] = 'ЕГРЮЛ'
    sheet_obj['F1'] = 'СМП'
    obj.create_sheet('ID')
    sheet_obj1 = obj['ID']
    sheet_obj1['A1'] = '4uiZ94AOro0g88uo6O5c'
    obj.save('Исход.xlsx')

TOKEN = 'необходимо вставить ваш токен зарегестрированного бота'
bot = telebot.TeleBot(TOKEN)


def keyboard_help(tel):
    markup_menu = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    if tel:
        btn_help = types.KeyboardButton('/start')
    else:
        btn_help = types.KeyboardButton('/start', request_contact=True)
    markup_menu.add(btn_help)
    return markup_menu


def check_id(user_id):  # проверяет есть ли id в списке допуска возвращает True и имя листа в котором находится этот
    # id, если нет в списках возвращает False, также возвращает True если телефон указан
    a = False
    b = False
    obj_ok = openpyxl.load_workbook('Исход.xlsx')
    sheet_obj_ok = obj_ok['ID']
    quantity = sheet_obj_ok.max_row
    bool_and_name = [a]
    for j in range(quantity):
        if (str(sheet_obj_ok.cell(row=(j + 1), column=1).value)).strip() == str(user_id):
            a = True
            if sheet_obj_ok.cell(row=(j + 1), column=5).value is not None:
                b = True
            bool_and_name.pop(0)
            bool_and_name.append(a)
            bool_and_name.append('ID')
            bool_and_name.append(b)  # проверяет есть ли номер в списке
            break

    return bool_and_name


def record_data_user(message):  # запись id в список разрешенныых
    a = False
    obj_ok = openpyxl.load_workbook('Исход.xlsx')
    sheets = obj_ok.sheetnames
    for i in sheets:
        sheet_obj_ok = obj_ok[i]
        quantity = sheet_obj_ok.max_row
        if (str(sheet_obj_ok.cell(row=1, column=1).value)).strip() == (str(message.text)).strip():
            sheet_obj_ok['A' + str(quantity + 1)] = str(message.chat.id)
            sheet_obj_ok['B' + str(quantity + 1)] = str(message.chat.first_name)
            sheet_obj_ok['C' + str(quantity + 1)] = str(message.chat.last_name)
            sheet_obj_ok['D' + str(quantity + 1)] = str(message.chat.username)
            obj_ok.save('Исход.xlsx')
            a = True
            break
    return a


@bot.message_handler(commands=['start'])
def send_welcom(message):
    user_id = message.chat.id
    res = check_id(user_id)
    bot.send_message(-434089215, str(user_id) + ' нажал start') # -434089215 это ID группы куда пересылаются сообщения
    if res[0]:
        bot.send_message(message.chat.id, "Здравствуйте! Я бот, который будет оповещать вас о окончании действия "
                                          "документов вторых частей а также ЭЦП.\n Необходимо добавить номер "
                                          " телефона.\n")
    else:
        bot.send_message(message.chat.id, "Здравствуйте! Я бот помощник.\nК сожалению Вас нет"
                                          " в базе, Вам необходимо ввести код доступа")


@bot.message_handler(content_types=['text'])
def echo(message):
    user_id = message.chat.id
    bot.send_message(-434089215, str(user_id) + ' прислал "' + message.text + '"')
    res = check_id(user_id)
    tex = message.text.strip().lower()
    if tex == 'данныеартем':
        doc = open('Исход.xlsx', 'rb')
        bot.send_document(message.chat.id, doc)  # отправка
        doc.close()
    if res[0]:  # сюда всю логику писаь надо
        markup_menu = keyboard_help(res[2])
        if res[2]:  # сюда писать логику
            if tex == 'хуй' or tex == 'пизда' or tex == 'пошел нахуй' or tex == 'сука' or tex == 'блядь' or tex == 'пидорас':
                bot.send_message(message.chat.id, 'Попрошу без выражений')
            elif tex == 'артем':
                bot.send_message(message.chat.id, '*СОЗДАТЕЛЬ* ЭТОГО БОТА', parse_mode="markdown")
            elif tex == 'оля':
                ola = 'CAACAgIAAxkBAAPuX0lYDQzXv-Cbl8nj5pd2sVlMaWYAAn8CAAJSFOEKojKL1C575mEbBA'
                bot.send_sticker(user_id, ola)
            elif tex == 'тамик':
                tam = 'CAACAgQAAxkBAAP4X0la-P1B0Kb0zfFjf4AwMT42beUAAmEBAAJwq9UEhlHspFDERSwbBA'
                bot.send_sticker(user_id, tam)
            elif tex == 'макс':
                ma = 'CAACAgIAAxkBAAIBCV9JXWHpTuYth5nOP87WjCc2N9vXAAItAwACtXHaBlJdSDo4DpaAGwQ'
                bot.send_sticker(user_id, ma)
            elif tex == 'амир':
                am = 'CAACAgUAAxkBAAIBFV9JXtUSEIQqOX8-6328pGgMmKN4AAItAAO8DacEbpHLQdiZ9LAbBA'
                bot.send_sticker(user_id, am)
            elif tex == 'армен':
                ar = 'CAACAgQAAxkBAAIBKF9JYPU7mpLLCLvoMh5yxMW7xi0aAAJEAwACNbs2AAGdm7TG8GUpshsE'
                bot.send_sticker(user_id, ar)
            elif tex == 'саня':
                sa = 'CAACAgIAAxkBAAIBLF9JYmQvcBjXgRY0H2Ye5lvrAS2QAAKGBQAC-gu2CGrh_7yWsoa8GwQ'
                bot.send_sticker(user_id, sa)
            elif tex == 'работать':
                bot.send_message(message.chat.id, 'Как меня все это заебало, ', parse_mode="markdown")
            elif tex == 'организации':
                organiza(user_id)
            elif tex == 'данные77':
                doc = open('Исход.xlsx', 'rb')
                bot.send_document(message.chat.id, doc)  # отправка
                doc.close()
            elif len(tex.split())==2 and tex.split()[-1] == 'получить':
                pass
            elif tex == 'получить':
                p = str(os.path.abspath('Исход.xlsx'))
                bot.send_message(user_id, p)
            else:
                find_org(user_id, tex)
        else:
            bot.send_message(message.chat.id, 'Вы не предоставили номер телефона.\nНажмите на клавишу start для '
                                              'предоставления номера телефона', reply_markup=markup_menu)

    else:
        kod_dostup = record_data_user(message)
        if kod_dostup:
            user_id = message.chat.id
            res = check_id(user_id)
            markup_menu = keyboard_help(res[2])
            bot.send_message(message.chat.id, "Вы успешно авторизовались в данном боте")
            putin = "CAACAgIAAxkBAAMbX0ap91v-6UhZAT6qM5pqSTYIKuYAAlcJAAJ5XOIJKyQBwA8ZVV4bBA"
            bot.send_sticker(user_id, putin)
            bot.send_message(message.chat.id, "Для дальнейшей работы"
                                              " нажмите клавишу start в приложении", reply_markup=markup_menu)

        else:
            bot.send_message(message.chat.id, "К сожалению данный код доступа не действителен. Пожалуйста уточните"
                                              " код доступа и повторите попытку аторизации")


@bot.message_handler(content_types=['contact'])  # получение номера и запись номера
def contacts(message):
    user_id = message.chat.id
    bot.send_message(-434089215, 'номер ' + str(message.contact.phone_number) + ' прислал ' + str(message.chat.id))
    res = check_id(user_id)
    al = True
    markup_menu = keyboard_help(al)
    obj_ok = openpyxl.load_workbook('Исход.xlsx')
    sheets = obj_ok.sheetnames
    for i in sheets:
        sheet_obj_ok = obj_ok[i]
        quantity = sheet_obj_ok.max_row
        for j in range(quantity):
            if (str(sheet_obj_ok.cell(row=(j + 1), column=1).value)).strip() == str(message.chat.id):
                sheet_obj_ok['E' + str(j + 1)] = str(message.contact.phone_number)
                obj_ok.save('Исход.xlsx')
                a = telebot.types.ReplyKeyboardRemove()
                bot.send_message(message.chat.id, "Благодарю за препдоставление номера телефона.\n"
                                                  " Теперь Вы будете получать сообщения об окончаниии действия "
                                                  "документов и ЭЦП", reply_markup=a)
                t = "CgACAgQAAxkBAAM1X0auxdxOAcCFXbwGmAnmB9P7PtMAAnMCAAKpgsVRzh1rI8qL2sUbBA"
                bot.send_document(user_id, t)
                bot.send_message(message.chat.id, 'Для ознакомления со списком организаций введите *ОРГАНИЗАЦИИ*', parse_mode="markdown")
                bot.send_message(message.chat.id, 'Для ознакомления с данными организации введите её наименование',
                                 parse_mode="markdown")

@bot.message_handler(content_types=['photo'])  # получение и отправка фото
def photo(message):
    pass
    # bot.send_message(message.chat.id, 'Такой команды нет.\n Нажмите на команду help для ознакомления с моими функциями')
    bot.send_photo(-434089215, message.photo[-1].file_id, caption='Прислал ' + str(message.chat.id))
    """ отправка документа и скачивание присланного фота
    doc = open('C:\\Users\\PC\\PycharmProjects\\telega\\ID.xlsx', 'rb')
    bot.send_document(message.chat.id, doc)  # отправка
    doc.close()
    file_info = bot.get_file(message.photo[-1].file_id)
    urllib.request.urlretrieve(f'http://api.telegram.org/file/bot{TOKEN}/{file_info.file_path}',
                               file_info.file_path)
    """


@bot.message_handler(content_types=['document'])  # получение и отправка документа
def docum(message):
    # bot.send_message(message.chat.id, 'Такой команды нет.\n Нажмите на команду help для ознакомления с моими функциями')
    bot.send_document(-434089215, message.document.file_id, caption='Прислал ' + str(message.chat.id))
    file_info = bot.get_file(message.document.file_id) #скачивание файла
    na = str(message.document.file_name).split('.')
    if na[0].strip() == 'Ис' and  na[1].strip() == 'xlsx':
        date_object = datetime.today()
        ti = str(date_object).split()[1].split(':')
        if not os.path.isdir("documents"):
            os.mkdir("documents")
        if not os.path.isdir("старое"):
            os.mkdir("старое")
        os.replace("Исход.xlsx", "старое/Исход" + ti[0] + ";" + ti[1] + ".xlsx")

        urllib.request.urlretrieve(f'http://api.telegram.org/file/bot{TOKEN}/{file_info.file_path}', file_info.file_path)
        fil = os.listdir("documents")[0]
        os.rename("documents/"+fil, "documents/Исход.xlsx")
        os.replace("documents/Исход.xlsx", "Исход.xlsx")
        bot.send_message(message.chat.id, 'Данные о сроках изменены')
@bot.message_handler(content_types=['sticker'])  # получение и отправка стикера
def stik(message):
    # bot.send_message(message.chat.id, 'Такой команды нет.\n Нажмите на команду help для ознакомления с моими функциями')
    bot.send_message(-434089215, 'Нижний стикер прислал ' + str(message.chat.id))
    bot.send_sticker(-434089215, message.sticker.file_id)


@bot.message_handler(content_types=['voice'])
def voi(message):
    # bot.send_message(message.chat.id, 'Такой команды нет.\n Нажмите на команду help для ознакомления с моими функциями')
    bot.send_voice(-434089215, message.voice.file_id, caption='Прислал ' + str(message.chat.id))


@bot.message_handler(content_types=['audio'])
def aud(message):
    # bot.send_message(message.chat.id, 'Такой команды нет.\n Нажмите на команду help для ознакомления с моими функциями')
    bot.send_audio(-434089215, message.audio.file_id, caption='Прислал ' + str(message.chat.id))


@bot.message_handler(content_types=['video'])
def vid(message):
    # bot.send_message(message.chat.id, 'Такой команды нет.\n Нажмите на команду help для ознакомления с моими функциями')
    bot.send_video(-434089215, message.video.file_id, caption='Прислал ' + str(message.chat.id))


@bot.message_handler(content_types=['video_note'])
def v_not(message):
    # bot.send_message(message.chat.id, 'Такой команды нет.\n Нажмите на команду help для ознакомления с моими функциями')
    bot.send_message(-434089215, 'Нижнее видеосообщение прислал ' + str(message.chat.id))
    # bot.send_video_note(-436648822, message.video_note.file_id)


@bot.message_handler(content_types=['animation'])
def v_not(message):
    # bot.send_message(message.chat.id, 'Такой команды нет.\n Нажмите на команду help для ознакомления с моими функциями')
    bot.send_message(-434089215, 'Нижнюю анимацию прислал ' + str(message.chat.id))
    # bot.send_animation(-436648822, message.animation.file_id, caption='Прислал ' + str(message.chat.id))

def obrabotka():
    while True:
        try:
            date_object = datetime.today()
            ti = str(date_object).split()[1].split(':')[0]
            date = str(date_object).split()[0].split('-')
            d2 = int(date[-1].strip())
            m2 = int(date[1].strip())
            y2 = int(date[0].strip())
            if int(ti) <= 20 and int(ti) > 6:
                obj_ok = openpyxl.load_workbook('Исход.xlsx')
                sheet_obj_ok = obj_ok['Данные']
                quantity = sheet_obj_ok.max_row

                for j in range(6):
                    if j + 1 == 1:
                        continue
                    else:
                        for i in range(quantity):
                            if i + 1 == 1:
                                continue
                            elif sheet_obj_ok.cell(row=i + 1, column=j + 1).value is None:
                                continue
                            else:
                                dat = str(sheet_obj_ok.cell(row=i + 1, column=j + 1).value).split()[0].split('-')
                                d1 = int(dat[-1])
                                m = int(dat[1])
                                y = int(dat[0])
                                q, p = srok(j + 1)
                                if m + q <= 12:
                                    m1 = m + q
                                    y1 = y
                                else:
                                    m1 = (m + q) % 12
                                    y1 = y + 1
                                try:
                                    t1 = datetime(year=y2, month=m2, day=d2, hour=1, minute=0, second=0)
                                except:
                                    t1 = datetime(year=y2, month=m2, day=28, hour=1, minute=0, second=0)
                                try:
                                    t2 = datetime(year=y1, month=m1, day=d1, hour=1, minute=0, second=0)
                                except:
                                    t2 = datetime(year=y1, month=m1, day=28, hour=1, minute=0, second=0)
                                t = int(str(t2 - t1).split()[0])
                                if t <= p:
                                    obj_ok1 = openpyxl.load_workbook('Исход.xlsx')
                                    sheet_obj_ok1 = obj_ok1['ID']
                                    quantity1 = sheet_obj_ok1.max_row
                                    for zz in range(quantity1):
                                        if zz + 1 == 1:
                                            continue
                                        elif sheet_obj_ok1.cell(row=zz + 1, column=5).value is None:
                                            continue
                                        else:
                                            iid = sheet_obj_ok1.cell(row=zz + 1, column=1).value
                                            firm = str(sheet_obj_ok.cell(row=i + 1, column=1).value)
                                            dok = str(sheet_obj_ok.cell(row=1, column=j + 1).value)
                                            try:
                                                bot.send_message(iid,
                                                                 'Истекает срок действия *' + dok + '* организации *' + firm + '*.\n'
                                                                                                                               'Ориентировочно осталось *' + str(
                                                                     t) + '* дней', parse_mode="markdown")
                                            except:
                                                pass
        except Exception as e:
            print(e)
            print('Ошибочка в рассылке сроков')
        time.sleep(21000)

task = threading.Thread(target=obrabotka)
task.start()


bot.polling(none_stop=True, interval=0)